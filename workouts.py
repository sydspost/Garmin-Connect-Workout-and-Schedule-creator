from openpyxl import load_workbook
from helium import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
import argparse
import os
import time

# Argument Parser
argParser = argparse.ArgumentParser(description='turbo Workout generator')
argParser.add_argument('--XLSX_filename', help='Specify input XLSX filename, default is "workoutschedule.xlsx"', default='workoutschedule.xlsx', dest='input')
argParser.add_argument('--output_dir', help='Specify output dir for JSON workout files, default is current directory', default=os.getcwd(), dest='output')
argParser.add_argument('-v', '--verbose', action='store_true')
argParser.add_argument('-u', '--username', help='Garmin Connect username', default="sydspost@gmail.com")
argParser.add_argument('-p', '--password', help='Garmin Connect password', default="X!do2019")
argParser.add_argument('-n', '--noschedule', help='Don \'t add workouts to calendar')
argParser.add_argument('-w', '--noworkout', help='Don \'t add workouts to Garmin Connect (implicit --n --noschedule)')
args = argParser.parse_args()

# Add Garmin Workout extenstion to Chrome
options = webdriver.ChromeOptions()
options.add_extension('./1.0.5_0.crx')			# Keep crx file in same directory as workouts.py !, or change path

# Login to Garmin Connect
if (not args.noworkout):
    driver=start_chrome(url='https://connect.garmin.com/modern/workouts', options=options)
    click('Accepteren')							# Accept cookies 
    write(args.username, into='E-mail')
    write(args.password, into='Wachtwoord')
    press(ENTER)

# Open Excel doc (default: workoutschedule.xlsx)
data = ""
wb = load_workbook(args.input)

# Worksheet Schedule
try:
    wsSchedule = wb["Schedule"]
except:
    print("Obligatory worksheet 'Schedule' is missing in Excel workbook")

# Worksheet Workout
try:
    wsWorkout = wb["Workout"]
except:
    print("Obligatory worksheet 'Workout' is missing in Excel workbook")

# Worksheet WorkoutType
try:
    wsWorkoutType = wb["workoutType"]
except:
    print("Obligatory worksheet 'workoutType' is missing in Excel workbook")

def defWorkout(workoutName, description):	# Create workout
    global data
    sportType = ""
    sportTypeId = ""
    sportTypeKey = ""
    workout = ""
    element = 0
    multisport = []
    multiSportWorkout = ""
    segmentOrder = 0
   
    # Build list of workouts from Worksheet Workout, column A
    workoutList = []
    for w in wsWorkout['A']:
        if (w.row > 1):
            workoutList.append(w.value)

    if (workoutName in workoutList):
        # workoutName & description
        data = "{\n"
        data += """       "workoutId": null,\n"""
        data += """       "ownerId": null,\n"""
        data += """       "workoutName": """ + "\"" + workoutName + "\",\n"
        if (description != None):
            data += """       "description": """ + "\"" + description.replace("\n", " ") + "\",\n"

        # Lookup SportType
        (sportTypeId, sportTypeKey) = lookupSportType(wsWorkout.cell(workoutList.index(workoutName)+2, 2).value)
        if (sportTypeId == "999"): # Multisport
            # sportType
            data += """       "sportType": {\n"""
            data += """               "sportTypeId": """ + sportTypeId + ",\n"
            data += """               "sportTypeKey": """ + "\"" + sportTypeKey + "\",\n"
            data += """               "displayOrder": 1\n"""
            data += "       },\n"
            
            workout = wsWorkout.cell(workoutList.index(workoutName)+2, 3).value.replace(" ", "")
            while (element < len(workout)):
                if (workout[element] == "|"):
                    sportType = workout[element+1:workout.find("|", element+1)]
                    element = workout.find("|", element+1)+1
                    if (workout.find("|", element) != -1):
                        multiSportWorkout = workout[element:workout.find("|", element)] + "#"
                        element = workout.find("|", element)-1
                    else:
                        multiSportWorkout = workout[element:len(workout)] + "#"
                        element = len(workout)
                    (sportTypeId, sportTypeKey) = lookupSportType(sportType)
                    multisport.append([multiSportWorkout, sportTypeId, sportTypeKey])
                element += 1

        else:
            # Lookup SportType
            (sportTypeId, sportTypeKey) = lookupSportType(wsWorkout.cell(workoutList.index(workoutName)+2, 2).value)
            multisport.append([wsWorkout.cell(workoutList.index(workoutName)+2, 3).value.replace(" ", "") + "#", sportTypeId, sportTypeKey])

            # sportType
            data += """       "sportType": {\n"""
            data += """               "sportTypeId": """ + multisport[0][1] + ",\n"
            data += """               "sportTypeKey": """ + "\"" + multisport[0][2] + "\",\n"
            data += """               "displayOrder": 1\n"""
            data += "       },\n"

        for w in multisport:
            segmentOrder += 1
            # workoutSegments
            data += """       "workoutSegments": [{\n"""
            data += """               "segmentOrder": """ + str(segmentOrder) + ",\n"
            data += """               "sportType": {\n"""
            data += """                       "sportTypeId": """ + w[1] + ",\n"
            data += """                       "sportTypeKey": """ + "\"" + w[2] + "\",\n"
            data += """                       "displayOrder": 1\n"""
            data += "               },\n"
            data += """               "workoutSteps": [{\n"""
             
            defParseWorkout(w[0], w[1])
 
def defParseWorkout(workout, sportTypeId):	# Build body of workout JSON file
    global data
    element = 0
    distance = ""
    time = ""
    zone = ""
    repeat = ""
    stepType = ""
    strokeType = ""
    rpmOrbpm = ""
    rpm = 0
    bpm = 0
    stepOrder = 0
    childStepId = 0
    numberOfIterations = 0
    workoutTargetTypeId = "1"
    workoutTargetTypeKey = "no.target"
    strokeTypeId = "0"
    strokeTypeKey = ""
    distanceOrtimeOrzoneOrrepeat = ""  

    # Walk through workout Description
    while (element < len(workout)):
        match workout[element]:
            case '0'| '1' | '2' | '3' | '4' | '5' | '6' | '7' | '8' | '9':
                distanceOrtimeOrzoneOrrepeat = distanceOrtimeOrzoneOrrepeat + workout[element]
            case 'm':					# OK, it's a distance
                distance = distanceOrtimeOrzoneOrrepeat
                time = ""
                distanceOrtimeOrzoneOrrepeat = ""
            case ':':					# OK, it's a time
                time = distanceOrtimeOrzoneOrrepeat + ":" + workout[element + 1] + workout[element + 2]
                element += 2
                distance = ""
                distanceOrtimeOrzoneOrrepeat = ""
            case 'Z':					# OK, it's a zone
                zone = workout[element + 1]
                element += 1
                distanceOrtimeOrzoneOrrepeat = ""
            case '*':					# OK, it's a repeat
                element += 1
                while (element < len(workout) and workout[element] != "+"):
                    repeat  += workout[element]
                    element += 1
            case '!':					# OK, it's a step- or stroketype
                element += 1
                while (element < len(workout) and workout[element] != "+" and workout[element] != "@"
                       and workout[element] != ")" and workout[element] != "#"):
                    if (sportTypeId != "4"):	# If sportTypeId <> Swimming, then it's a stepType
                        stepType += workout[element]
                    else:						# else it's a strokeType
                        strokeType += workout[element]
                        if (strokeType == "RUST"):
                            stepType = strokeType
                            strokeType = ""
                    element += 1
                element -= 1
            case '@':					# OK, it's a Zone, RPM or BPM, Zone is ignored because it's already accessed under 'Z'
                if (workout[element + 1] != "Z"):
                    element += 1
                    while (element < len(workout) and workout[element] != "+" and workout[element] != "#"):
                        rpmOrbpm += workout[element]
                        element += 1
                    match rpmOrbpm[-3:]:
                        case "rpm":		# It's RPM
                            rpm = int(rpmOrbpm[:len(rpmOrbpm)-3])
                        case "bpm":		# It's BPM
                            bpm = int(rpmOrbpm[:len(rpmOrbpm)-3])
                            
                    rpmOrbpm = ""                    
                    element -= 1
            case '(':					# It's the start of a repeatgroup
                childStepId += 1
                stepOrder += 1
                
                # find number of iterations of this repeatgroup
                numberOfIterations = workout[workout.find('*', element)+1:workout.find('+', workout.find('*', element))]

                data += """                   "type": "RepeatGroupDTO",\n"""
                data += """                   "stepOrder": """ + str(stepOrder) + ",\n"
                data += """                   "stepType": {\n"""
                data += """                           "stepTypeId": 6,\n"""
                data += """                           "stepTypeKey": "repeat",\n"""
                data += """                           "displayOrder": 6\n"""
                data += """                   },\n"""
                data += """                   "childStepId": """ + str(childStepId) + ",\n"
                data += """                   "numberOfIterations": """ + str(numberOfIterations) + ",\n"
                data += """                   "workoutSteps": [{\n"""
            case '+' | ')' | '#':		# Next workoutstep, so finish current step
                if (stepType != ""):
                    # find stepTypeId, stepTypeKey and stepTypeDescription in worksheet workoutType
                    (stepTypeId, stepTypeKey, stepTypeDescription) = lookupStepType(stepType)
                    if (stepTypeId == ""):
                        stepTypeId = "null"
                        stepTypeKey = "null"
                        stepTypeDescription = "null"
                else:					# if empty, then it's a Interval steptype
                    (stepTypeId, stepTypeKey, stepTypeDescription) = lookupStepType("INT")
                
                if (stepTypeId == 5):	# Rest
                    conditionTypeId = "1"
                    conditionTypeKey = "lap.button"
                    endConditionValue = "0.0"

                if (distance != ""):	# Distance or Time ?
                    conditionTypeId = "3"
                    conditionTypeKey = "distance"
                    endConditionValue = '{0:.1f}'.format(float(distance))                 
                else:					# Time
                    conditionTypeId = "2"
                    conditionTypeKey = "time"
                    t=int(time[0:2])*60 + int(time[3:5])
                    endConditionValue = '{:.1f}'.format(float(t))
                    
                if (stepType == "RUST"):# Rest
                    conditionTypeId = "8"
                    conditionTypeKey = "fixed.rest"
                    t=int(time[0:2])*60 + int(time[3:5])
                    endConditionValue = '{:.1f}'.format(float(t))
                    
                stepOrder += 1
                
                data += """                       "type": "ExecutableStepDTO",\n"""
                data += """                       "stepId": null,\n"""
                data += """                       "stepOrder": """ + str(stepOrder) + ",\n"""
                data += """                       "stepType": {\n"""
                data += """                               "stepTypeId": """ + stepTypeId + ",\n"
                data += """                               "stepTypeKey": """ + stepTypeKey + ",\n"
                data += """                               "displayOrder": 3\n"""
                data += """                       },\n"""
                data += """                       "childStepId": """ + str(childStepId) + ",\n"
                data += """                       "description": """ + stepTypeDescription + ",\n"
                data += """                       "endCondition": {\n"""
                data += """                               "conditionTypeId": """ + conditionTypeId + ",\n"
                data += """                               "conditionTypeKey": """ + "\"" + conditionTypeKey + "\",\n"
                data += """                               "displayOrder": 2,\n"""
                data += """                               "displayable": true\n"""
                data += """                       },\n"""
                data += """                       "endConditionValue": """ + endConditionValue + ",\n"
                
                if (conditionTypeId == "3"):	# Distance
                    data += """                       "preferredEndConditionUnit": {\n"""
                    data += """                               "unitId": 1,\n"""
                    data += """                               "unitKey": "meter",\n"""
                    data += """                               "factor": 100.0\n"""
                    data += """                       },\n"""
                    
                if (zone != ""):				# Zone detected, so workoutTargetType should be heart.rate.zone
                    workoutTargetTypeId = "4"
                    workoutTargetTypeKey = "heart.rate.zone"
                    
                data += """                       "targetType": {\n"""
                data += """                               "workoutTargetTypeId": """ + workoutTargetTypeId + ",\n"
                data += """                               "workoutTargetTypeKey": """ + "\"" + workoutTargetTypeKey + "\",\n"
                data += """                               "displayOrder": 1\n"""
                data += """                       },\n"""
                if (zone != ""):
                    data += """                       "zoneNumber": """ + zone + ",\n"
                    workoutTargetTypeId = "1"
                    workoutTargetTypeKey = "no.target"
                
                if (rpm != 0 or bpm != 0):		# RPM or BPM, so fill secondaryTargetType when Zone is detected
                    if (zone != ""):
                        data += """                       "secondaryTargetType": {\n"""
                        if (rpm != 0):
                            data += """                               "workoutTargetTypeId": 3,\n"""
                            data += """                               "workoutTargetTypeKey": "cadence",\n"""
                            data += """                               "displayOrder": 3\n"""
                            data += """                       },\n"""
                            data += """                       "secondaryTargetValueOne": """ + str(rpm - 10) + ",\n"
                            data += """                       "secondaryTargetValueTwo": """ + str(rpm + 10) + ",\n"
                        else:
                            data += """                               "workoutTargetTypeId": 4,\n"""
                            data += """                               "workoutTargetTypeKey": "heart.rate.zone",\n"""
                            data += """                               "displayOrder": 3\n"""
                            data += """                       },\n"""
                            data += """                       "secondaryTargetValueOne": """ + str(bpm - 10) + ",\n"
                            data += """                       "secondaryTargetValueTwo": """ + str(bpm + 10) + ",\n"
                    else:						# no Zone detected, so fill workoutTargetType with RPM of BPM
                        data += """                       "targetType": {\n"""
                        if (rpm != 0):
                            data += """                               "workoutTargetTypeId": 3,\n"""
                            data += """                               "workoutTargetTypeKey": "cadence",\n"""
                            data += """                               "displayOrder": 3\n"""
                            data += """                       },\n"""
                            data += """                       "targetValueOne": """ + str(rpm - 10) + ",\n"
                            data += """                       "targetValueTwo": """ + str(rpm + 10) + ",\n"
                        else:
                            data += """                               "workoutTargetTypeId": 4,\n"""
                            data += """                               "workoutTargetTypeKey": "heart.rate.zone",\n"""
                            data += """                               "displayOrder": 3\n"""
                            data += """                       },\n"""
                            data += """                       "targetValueOne": """ + str(bpm - 10) + ",\n"
                            data += """                       "targetValueTwo": """ + str(bpm + 10) + ",\n"

                    rpm = 0
                    bpm = 0
                   
 
                if (sportTypeId == "4" and stepTypeId != "5"):	# Swimming, but no rest, fill StrokeType
                    (strokeTypeId, strokeTypeKey) = lookupStrokeType(strokeType)
                    data += """                       "strokeType": {\n"""
                    data += """                               "strokeTypeId": """ + strokeTypeId + ",\n"
                    data += """                               "strokeTypeKey": """ + "\"" + strokeTypeKey + "\",\n"
                    data += """                               "displayOrder": 1\n"""
                    data += """                       },\n"""
                    
                data += """                       "equipmentType": {\n"""
                data += """                               "equipmentTypeId": 0,\n"""
                data += """                               "displayOrder": 0\n"""
                data += """                       }\n"""

                if (workout[element] == '+'):					# Finish for next Workoutstep
                    if (sportTypeId == "4" and numberOfIterations == 0):
                        stepOrder += 1
                        data += """               }, {\n"""
                        data += """               "type": "ExecutableStepDTO",\n"""
                        data += """               "stepId": null,\n"""
                        data += """               "stepOrder": """ + str(stepOrder) + ",\n"
                        data += """               "stepType": {\n"""
                        data += """                       "stepTypeId": 5,\n"""
                        data += """                       "stepTypeKey": "rest",\n"""
                        data += """                       "displayOrder": 5\n"""
                        data += """               },\n"""
                        data += """               "childStepId": null,\n"""
                        data += """               "description": null,\n"""
                        data += """               "endCondition": {\n"""
                        data += """                       "conditionTypeId": 1,\n"""
                        data += """                       "conditionTypeKey": "lap.button",\n"""
                        data += """                       "displayOrder": 1,\n"""
                        data += """                       "displayable": true\n"""
                        data += """               }\n"""
                    data += """               }, {\n"""
                
                # Reset variables for next Workout step
                stepType = ""
                strokeType = ""
                stepTypeId = ""
                stepTypeKey = ""
                stepTypeDescription = ""
                conditionTypeId = ""
                zone = ""
                strokeTypeId = "0"
                strokeTypeKey = ""
                
                # Finish repeatgroup
                if (workout[element] == ')'):
                    data += """                   }],\n"""
                    data += """                   "endConditionValue": """ + '{0:.1f}'.format(float(numberOfIterations)) + ",\n"
                    data += """                   "endCondition": {\n"""
                    data += """                           "conditionTypeId": 7,\n"""
                    data += """                           "conditionTypeKey": "iterations",\n"""
                    data += """                           "displayOrder": 7,\n"""
                    data += """                           "displayable": false\n"""
                    data += """                   },\n"""
                    data += """                   "smartRepeat": false\n"""
                    data += """               }, {\n"""
                    numberOfIterations = 0
                    
                    if (sportTypeId == "4"):	# Swimming
                        stepOrder += 1
                        data += """               "type": "ExecutableStepDTO",\n"""
                        data += """               "stepId": null,\n"""
                        data += """               "stepOrder": """ + str(stepOrder) + ",\n"
                        data += """               "stepType": {\n"""
                        data += """                       "stepTypeId": 5,\n"""
                        data += """                       "stepTypeKey": "rest",\n"""
                        data += """                       "displayOrder": 5\n"""
                        data += """               },\n"""
                        data += """               "childStepId": null,\n"""
                        data += """               "description": null,\n"""
                        data += """               "endCondition": {\n"""
                        data += """                       "conditionTypeId": 1,\n"""
                        data += """                       "conditionTypeKey": "lap.button",\n"""
                        data += """                       "displayOrder": 1,\n"""
                        data += """                       "displayable": true\n"""
                        data += """               }\n"""
                        data += """               }, {\n"""
                        
                    element = workout.find("+", element)

                    
        element += 1

    # workoutSegments
    data += "                }]\n"
    data += "        }],\n"
    data += """        "poolLength": 25.0,\n"""
    data += """        "poolLengthUnit": {\n"""
    data += """                "unitId": 1,\n"""
    data += """                "unitKey": "meter",\n"""
    data += """                "factor": 100.0\n"""
    data += "        },\n"
    data += """        "avgTrainingSpeed": 0.0,\n"""
    data += """        "estimatedDistanceUnit": {\n"""
    data += """                "unitId": null,\n"""
    data += """                "unitKey": null,\n"""
    data += """                "factor": null\n"""
    data += """        },\n"""
    data += """        "shared": false\n"""
    data += "}\n"

def lookupSportType(sportType):				# Lookup SportType in worksheet workoutType
    sportTypeId = ""
    sportTypeKey = ""

    for sporttypeRow in wsWorkoutType:
        if (sporttypeRow[2].value == "sportTypeId" and sporttypeRow[0].value == sportType):
            sportTypeId = str(sporttypeRow[1].value)
            sportTypeKey = sporttypeRow[3].value
            
    return (sportTypeId, sportTypeKey)
            
def lookupStepType(stepType):				# Lookup StepType in worksheet workoutType
    stepTypeId = ""
    stepTypeKey = ""
    stepTypeDescription = ""
    
    for stepTypeRow in wsWorkoutType:
        if (stepTypeRow[2].value == "stepTypeId" and stepTypeRow[0].value == stepType):
            stepTypeId = str(stepTypeRow[1].value)
            stepTypeKey = "\"" + stepTypeRow[3].value + "\""
            stepTypeDescription = "\"" + stepTypeRow[4].value + "\""
            
    return (stepTypeId, stepTypeKey, stepTypeDescription)

def lookupStrokeType(strokeType):			# Lookup StrokeType in worksheet workoutType
    strokeTypeId = ""
    strokeTypeKey = ""
    
    for strokeTypeRow in wsWorkoutType:
        if (strokeTypeRow[2].value == "swimStrokeType" and strokeTypeRow[0].value == strokeType):
            strokeTypeId = str(strokeTypeRow[1].value)
            strokeTypeKey = strokeTypeRow[3].value
            
    return (strokeTypeId, strokeTypeKey)


# Loop: Iterate thru Scheduled workouts
for row in range (2, wsSchedule.max_row+1):
    defWorkout(wsSchedule.cell(row=row, column=2).value, wsSchedule.cell(row=row, column=3).value)

    if (data != ""):
        if (args.verbose): print(data)
        f = open(args.output + "\\" + wsSchedule.cell(row=row, column=2).value + ".json", "w", encoding='utf-8')
        f.write(data)
        f.close()
        data=""
        
        # Import Workout into Garmin Connect
        if (not args.noworkout):
            wait_until(Text("Workouts").exists, 10)
            print(args.output + "\\" + wsSchedule.cell(row=row, column=2).value + ".json")
            driver.find_element(By.ID, 'garmin-share-workout-file').send_keys(args.output + "\\" + wsSchedule.cell(row=row, column=2).value + ".json")
            time.sleep(2)
            try:
                alert = driver.switch_to.alert
                alert.accept()
            except:
                pass
        
        # Assign workout to calendar
        if (not args.noschedule or not args.noworkout):
            wait_until(Text("Copyright").exists, 10)
            time.sleep(2)
            driver.find_element(By.LINK_TEXT, 'Voeg toe aan agenda').click()
            wait_until(Text("PLANNING").exists, 10)
            month=wsSchedule.cell(row=row, column=1).value.month
            if (month < datetime.now().month):
                for i in range(datetime.now().month - month):
                    click(S("//span[@id='workout-calendar-arrow-left']/a/i"))  # shift month before
            else:
                for i in range(month - datetime.now().month):
                    click(S("//span[@id='workout-calendar-arrow-right']/a/i")) # shift month after
            click(S("//td[@id='" + str(month-1) + "-" + str(wsSchedule.cell(row=row, column=1).value.day) + "']"))
        go_to('https://connect.garmin.com/modern/workouts')
        time.sleep(2)
        
kill_browser()
    
 